from datetime import date, timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.test_dashboard import create_center, create_project, create_user


def test_dashboard_v31_crud_warning_and_permissions(
    client: TestClient,
    admin_headers: dict[str, str],
) -> None:
    project_id = create_project(client, admin_headers, "V31")
    center_id = create_center(client, admin_headers, project_id, "V31")
    today = date.today()

    readonly_headers = create_user(
        client,
        admin_headers,
        "v31_readonly",
        "readonly",
        project_ids=[project_id],
    )
    forbidden = client.post(
        "/api/dashboard/v31/milestones",
        headers=readonly_headers,
        json={
            "project_id": project_id,
            "center_id": center_id,
            "milestone_name": "伦理批件",
        },
    )
    assert forbidden.status_code == 403

    create_response = client.post(
        "/api/dashboard/v31/milestones",
        headers=admin_headers,
        json={
            "project_id": project_id,
            "center_id": center_id,
            "milestone_name": "伦理批件",
            "planned_date": (today - timedelta(days=1)).isoformat(),
            "status": "in_progress",
            "owner": "贾飞",
        },
    )
    assert create_response.status_code == 201
    milestone_id = create_response.json()["id"]

    overview = client.get(
        f"/api/dashboard/v31/project/{project_id}/overview",
        headers=admin_headers,
    )
    assert overview.status_code == 200
    assert overview.json()["counts"]["milestones"] == 1
    assert overview.json()["deviation_warnings"][0]["warning_level"] == "overdue"

    update_response = client.patch(
        f"/api/dashboard/v31/milestones/{milestone_id}",
        headers=admin_headers,
        json={"status": "done", "actual_date": today.isoformat()},
    )
    assert update_response.status_code == 200
    assert update_response.json()["status"] == "done"

    cleared = client.get(
        f"/api/dashboard/v31/project/{project_id}/overview",
        headers=admin_headers,
    )
    assert cleared.status_code == 200
    assert cleared.json()["deviation_warnings"] == []

    delete_response = client.delete(
        f"/api/dashboard/v31/milestones/{milestone_id}",
        headers=admin_headers,
    )
    assert delete_response.status_code == 204


def test_dashboard_v31_import_export_upsert(
    client: TestClient,
    admin_headers: dict[str, str],
) -> None:
    project_id = create_project(client, admin_headers, "V31_IMPORT")
    create_center(client, admin_headers, project_id, "V31_IMPORT")

    template_response = client.get(
        "/api/dashboard/v31/import-template/important-tasks",
        headers=admin_headers,
    )
    assert template_response.status_code == 200
    workbook = load_workbook(BytesIO(template_response.content))
    worksheet = workbook.active
    worksheet.append(
        [
            "",
            "",
            "V3.1 数据看板验收",
            "贾飞",
            date.today().isoformat(),
            "",
            "open",
            "important",
            "urgent",
            "首轮导入",
        ]
    )
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    first_import = client.post(
        f"/api/dashboard/v31/import/important-tasks?project_id={project_id}",
        headers=admin_headers,
        files={
            "file": (
                "important-tasks.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert first_import.status_code == 200
    assert first_import.json()["created_count"] == 1

    second_import = client.post(
        f"/api/dashboard/v31/import/important-tasks?project_id={project_id}",
        headers=admin_headers,
        files={
            "file": (
                "important-tasks.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert second_import.status_code == 200
    assert second_import.json()["updated_count"] == 1

    export_response = client.get(
        f"/api/dashboard/v31/export/important-tasks?project_id={project_id}",
        headers=admin_headers,
    )
    assert export_response.status_code == 200
    exported = load_workbook(BytesIO(export_response.content)).active
    assert exported["C3"].value == "V3.1 数据看板验收"
