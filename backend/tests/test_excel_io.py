from collections.abc import Iterable
from contextlib import contextmanager
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.core.database import get_db
from app.models import SubjectItem


@contextmanager
def db_session(client: TestClient):
    override = client.app.dependency_overrides[get_db]
    session_generator = override()
    db = next(session_generator)
    try:
        yield db
    finally:
        session_generator.close()


def login_headers(client: TestClient, username: str, password: str) -> dict[str, str]:
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def role_id_by_name(client: TestClient, headers: dict[str, str], name: str) -> int:
    response = client.get("/api/roles", headers=headers)
    assert response.status_code == 200
    return next(role["id"] for role in response.json() if role["name"] == name)


def permission_id_by_code(client: TestClient, headers: dict[str, str], code: str) -> int:
    response = client.get("/api/permissions", headers=headers)
    assert response.status_code == 200
    return next(permission["id"] for permission in response.json() if permission["code"] == code)


def xlsx_bytes(headers: list[str], rows: Iterable[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def upload_xlsx(
    client: TestClient,
    headers: dict[str, str],
    path: str,
    file_content: bytes,
) -> dict:
    response = client.post(
        path,
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                file_content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    return response.json()


def create_project(client: TestClient, headers: dict[str, str], suffix: str) -> int:
    response = client.post(
        "/api/projects",
        headers=headers,
        json={
            "name": f"P7 项目 {suffix}",
            "code": f"P7_PROJECT_{suffix}",
            "description": "",
            "status": "active",
        },
    )
    assert response.status_code == 201
    return response.json()["id"]


def create_center(client: TestClient, headers: dict[str, str], project_id: int, suffix: str) -> int:
    response = client.post(
        "/api/centers",
        headers=headers,
        json={
            "project_id": project_id,
            "name": f"P7 中心 {suffix}",
            "code": f"P7_CENTER_{suffix}",
            "contact_person": "",
            "status": "active",
            "description": "",
        },
    )
    assert response.status_code == 201
    return response.json()["id"]


def create_stage(client: TestClient, headers: dict[str, str], project_id: int) -> int:
    response = client.post(
        "/api/stages",
        headers=headers,
        json={
            "project_id": project_id,
            "name": "启动阶段",
            "code": "STARTUP",
            "sort_order": 1,
            "description": "",
        },
    )
    assert response.status_code == 201
    return response.json()["id"]


def create_user(
    client: TestClient,
    admin_headers: dict[str, str],
    username: str,
    role: str,
    project_ids: list[int] | None = None,
    center_ids: list[int] | None = None,
) -> dict[str, str]:
    response = client.post(
        "/api/users",
        headers=admin_headers,
        json={
            "username": username,
            "password": "User@12345",
            "full_name": username,
            "email": None,
            "is_active": True,
            "role_ids": [role_id_by_name(client, admin_headers, role)],
            "project_ids": project_ids or [],
            "center_ids": center_ids or [],
        },
    )
    assert response.status_code == 201
    return login_headers(client, username, "User@12345")


def workbook_rows(response_content: bytes) -> list[list[object]]:
    workbook = load_workbook(BytesIO(response_content), data_only=True)
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def test_import_templates_and_four_import_upserts(
    client: TestClient,
    admin_headers: dict[str, str],
) -> None:
    template = client.get("/api/import/templates/projects", headers=admin_headers)
    assert template.status_code == 200
    rows = workbook_rows(template.content)
    assert rows[0][:2] == ["项目编码", "项目名称"]
    assert rows[1][:2] == ["code", "name"]
    subject_template = client.get("/api/import/templates/subjects", headers=admin_headers)
    assert subject_template.status_code == 200
    subject_template_rows = workbook_rows(subject_template.content)
    assert "知情时间" in subject_template_rows[0]
    assert "访视5日期" in subject_template_rows[0]

    project_result = upload_xlsx(
        client,
        admin_headers,
        "/api/import/projects",
        xlsx_bytes(
            ["code", "name", "status", "description"],
            [["P7_IMPORT_PROJECT", "P7 导入项目", "active", "首次导入"]],
        ),
    )
    assert project_result["created_count"] == 1
    assert project_result["updated_count"] == 0

    update_result = upload_xlsx(
        client,
        admin_headers,
        "/api/import/projects",
        xlsx_bytes(
            ["code", "name", "status", "description"],
            [["P7_IMPORT_PROJECT", "P7 导入项目更新", "active", "Upsert"]],
        ),
    )
    assert update_result["created_count"] == 0
    assert update_result["updated_count"] == 1

    project = client.get("/api/projects", headers=admin_headers).json()[0]
    assert project["name"] == "P7 导入项目更新"
    project_id = project["id"]

    center_result = upload_xlsx(
        client,
        admin_headers,
        "/api/import/centers",
        xlsx_bytes(
            ["project_code", "code", "name", "contact_person", "status", "description"],
            [["P7_IMPORT_PROJECT", "P7_IMPORT_CENTER", "P7 导入中心", "CRC", "active", ""]],
        ),
    )
    assert center_result["created_count"] == 1
    centers = client.get(f"/api/centers?project_id={project_id}", headers=admin_headers)
    center_id = centers.json()[0]["id"]

    stage_id = create_stage(client, admin_headers, project_id)
    template_result = upload_xlsx(
        client,
        admin_headers,
        "/api/import/stage-templates",
        xlsx_bytes(
            [
                "project_code",
                "stage_code",
                "item_code",
                "item_name",
                "required",
                "sort_order",
                "description",
            ],
            [["P7_IMPORT_PROJECT", "STARTUP", "ETHICS", "伦理批件", "是", 1, ""]],
        ),
    )
    assert template_result["created_count"] == 1
    templates = client.get(
        f"/api/stage-templates?project_id={project_id}&stage_id={stage_id}",
        headers=admin_headers,
    )
    assert templates.status_code == 200
    assert templates.json()[0]["item_code"] == "ETHICS"

    subject_result = upload_xlsx(
        client,
        admin_headers,
        "/api/import/subjects",
        xlsx_bytes(
            [
                "project_code",
                "center_code",
                "screening_no",
                "gender",
                "age",
                "enrolled_at",
                "informed_at",
                "visit1_date",
                "visit2_date",
                "visit3_date",
                "visit4_date",
                "visit5_date",
            ],
            [
                [
                    "P7_IMPORT_PROJECT",
                    "P7_IMPORT_CENTER",
                    "P7-S001",
                    "女",
                    38,
                    "2026-05-01",
                    "2026-05-01 09:30",
                    "2026-05-02",
                    "2026-05-03",
                    "2026-05-04",
                    "2026-05-05",
                    "2026-05-06",
                ]
            ],
        ),
    )
    assert subject_result["created_count"] == 1
    subject = client.get(
        f"/api/subjects?project_id={project_id}&center_id={center_id}",
        headers=admin_headers,
    ).json()[0]
    assert subject["screening_no"] == "P7-S001"
    assert subject["informed_at"].startswith("2026-05-01T09:30")
    assert subject["visit5_date"] == "2026-05-06"

    with db_session(client) as db:
        item_count = db.scalar(
            select(func.count(SubjectItem.id)).where(SubjectItem.subject_id == subject["id"])
        )
    assert item_count == 19


def test_import_validation_is_all_or_nothing(
    client: TestClient,
    admin_headers: dict[str, str],
) -> None:
    project_id = create_project(client, admin_headers, "VALIDATION")
    create_center(client, admin_headers, project_id, "VALID")
    result = upload_xlsx(
        client,
        admin_headers,
        "/api/import/subjects",
        xlsx_bytes(
            [
                "project_code",
                "center_code",
                "screening_no",
                "gender",
                "age",
                "enrolled_at",
                "informed_at",
            ],
            [
                [
                    "P7_PROJECT_VALIDATION",
                    "P7_CENTER_VALID",
                    "P7-V-001",
                    "男",
                    45,
                    "2026-05-02",
                    "2026-05-02 09:30",
                ],
                [
                    "P7_PROJECT_VALIDATION",
                    "MISSING_CENTER",
                    "P7-V-002",
                    "女",
                    "bad",
                    "2026-99-99",
                    "bad-time",
                ],
            ],
        ),
    )
    assert result["total_rows"] == 2
    assert result["created_count"] == 0
    assert {error["field"] for error in result["errors"]} == {
        "center_code",
        "age",
        "enrolled_at",
        "informed_at",
    }

    subjects = client.get(f"/api/subjects?project_id={project_id}", headers=admin_headers)
    assert subjects.status_code == 200
    assert subjects.json() == []


def test_import_and_export_permissions_and_scope(
    client: TestClient,
    admin_headers: dict[str, str],
) -> None:
    project_id = create_project(client, admin_headers, "SCOPE")
    center_a = create_center(client, admin_headers, project_id, "A")
    create_center(client, admin_headers, project_id, "B")
    scoped_headers = create_user(
        client,
        admin_headers,
        "p7_center_manager",
        "center_manager",
        center_ids=[center_a],
    )
    readonly_headers = create_user(
        client,
        admin_headers,
        "p7_readonly",
        "readonly",
        project_ids=[project_id],
    )

    denied_master_import = client.post(
        "/api/import/centers",
        headers=scoped_headers,
        files={
            "file": (
                "centers.xlsx",
                xlsx_bytes(
                    ["project_code", "code", "name"],
                    [["P7_PROJECT_SCOPE", "P7_CENTER_C", "P7 中心 C"]],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert denied_master_import.status_code == 403

    allowed = client.post(
        "/api/import/subjects",
        headers=scoped_headers,
        files={
            "file": (
                "subjects.xlsx",
                xlsx_bytes(
                    ["project_code", "center_code", "screening_no"],
                    [["P7_PROJECT_SCOPE", "P7_CENTER_A", "P7-A-001"]],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert allowed.status_code == 200
    assert allowed.json()["created_count"] == 1

    denied_scope = client.post(
        "/api/import/subjects",
        headers=scoped_headers,
        files={
            "file": (
                "subjects.xlsx",
                xlsx_bytes(
                    ["project_code", "center_code", "screening_no"],
                    [["P7_PROJECT_SCOPE", "P7_CENTER_B", "P7-B-001"]],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert denied_scope.status_code == 403

    denied_write = client.post(
        "/api/import/subjects",
        headers=readonly_headers,
        files={
            "file": (
                "subjects.xlsx",
                xlsx_bytes(
                    ["project_code", "center_code", "screening_no"],
                    [["P7_PROJECT_SCOPE", "P7_CENTER_A", "P7-A-002"]],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert denied_write.status_code == 403

    readonly_export = client.get(
        f"/api/export/project-progress?project_id={project_id}",
        headers=readonly_headers,
    )
    assert readonly_export.status_code == 200

    export_permission_id = permission_id_by_code(client, admin_headers, "master_data:read")
    role = client.post(
        "/api/roles",
        headers=admin_headers,
        json={
            "name": "p7_no_export",
            "label": "无导出",
            "description": "",
            "permission_ids": [export_permission_id],
        },
    )
    assert role.status_code == 201
    user = client.post(
        "/api/users",
        headers=admin_headers,
        json={
            "username": "p7_no_export",
            "password": "User@12345",
            "full_name": "无导出",
            "email": None,
            "is_active": True,
            "role_ids": [role.json()["id"]],
            "project_ids": [project_id],
            "center_ids": [],
        },
    )
    assert user.status_code == 201
    no_export_headers = login_headers(client, "p7_no_export", "User@12345")
    forbidden = client.get(
        f"/api/export/project-progress?project_id={project_id}",
        headers=no_export_headers,
    )
    assert forbidden.status_code == 403


def test_exports_include_expected_workbooks_and_missing_unmaterialized_stage_items(
    client: TestClient,
    admin_headers: dict[str, str],
) -> None:
    project_id = create_project(client, admin_headers, "EXPORT")
    center_id = create_center(client, admin_headers, project_id, "EXPORT")
    create_stage(client, admin_headers, project_id)
    upload_xlsx(
        client,
        admin_headers,
        "/api/import/stage-templates",
        xlsx_bytes(
            [
                "project_code",
                "stage_code",
                "item_code",
                "item_name",
                "required",
                "sort_order",
            ],
            [["P7_PROJECT_EXPORT", "STARTUP", "ETHICS", "伦理批件", "是", 1]],
        ),
    )
    upload_xlsx(
        client,
        admin_headers,
        "/api/import/subjects",
        xlsx_bytes(
            ["project_code", "center_code", "screening_no"],
            [["P7_PROJECT_EXPORT", "P7_CENTER_EXPORT", "P7-E-001"]],
        ),
    )

    progress = client.get(
        f"/api/export/project-progress?project_id={project_id}",
        headers=admin_headers,
    )
    assert progress.status_code == 200
    assert "project-progress.xlsx" in progress.headers["content-disposition"]
    assert workbook_rows(progress.content)[0][:5] == [
        "项目编码",
        "项目名称",
        "可见中心数",
        "受试者数",
        "完成案例数",
    ]

    for path in [
        "/api/export/center-status",
        "/api/export/subject-completeness",
        "/api/export/missing-items",
    ]:
        response = client.get(f"{path}?project_id={project_id}", headers=admin_headers)
        assert response.status_code == 200
        assert response.content.startswith(b"PK")

    subject_completeness = client.get(
        f"/api/export/subject-completeness?project_id={project_id}",
        headers=admin_headers,
    )
    subject_headers = workbook_rows(subject_completeness.content)[0]
    assert "知情时间" in subject_headers
    assert "访视5日期" in subject_headers

    missing = client.get(
        f"/api/export/missing-items?project_id={project_id}&center_id={center_id}",
        headers=admin_headers,
    )
    rows = workbook_rows(missing.content)
    row_types = [row[0] for row in rows[1:]]
    assert "阶段资料" in row_types
    assert "受试者数据项" in row_types
