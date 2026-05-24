from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import Select, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.deps import AccessContext
from app.models import (
    Center,
    DashboardClinicalEvent,
    DashboardDeviceHandover,
    DashboardDeviceIssue,
    DashboardEnrollmentPlan,
    DashboardImportantTask,
    DashboardMilestone,
    DashboardSubjectOverview,
    DashboardSubjectResult,
    Project,
    Subject,
)
from app.schemas.dashboard import DashboardV31ImportErrorRead, DashboardV31ImportResultRead

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DONE_STATUSES = {"done", "completed", "complete", "finished", "closed", "已完成", "完成", "已关闭"}


@dataclass(frozen=True)
class DashboardColumn:
    key: str
    label: str
    kind: str = "text"
    required: bool = False
    note: str = ""


@dataclass(frozen=True)
class DashboardKindConfig:
    kind: str
    path: str
    title: str
    model: type
    columns: list[DashboardColumn]
    unique_keys: list[str]


BASE_COLUMNS = [
    DashboardColumn("id", "ID", "int", False, "导入时填写 ID 可更新指定记录"),
    DashboardColumn("center_code", "中心编码", "text", False, "留空表示项目级记录"),
]


DASHBOARD_V31_CONFIGS: dict[str, DashboardKindConfig] = {
    "milestones": DashboardKindConfig(
        "milestones",
        "milestones",
        "进度甘特图",
        DashboardMilestone,
        BASE_COLUMNS
        + [
            DashboardColumn(
                "milestone_name",
                "里程碑",
                "text",
                True,
                "如伦理批件/合同完成/省局备案/启动时间/方案修正案/入组",
            ),
            DashboardColumn("planned_date", "计划日期", "date"),
            DashboardColumn("actual_date", "实际日期", "date"),
            DashboardColumn("status", "状态", "text", False, "not_started/in_progress/done"),
            DashboardColumn("owner", "负责人"),
            DashboardColumn("notes", "备注"),
        ],
        ["center_id", "milestone_name"],
    ),
    "enrollment-plans": DashboardKindConfig(
        "enrollment-plans",
        "enrollment-plans",
        "入组计划表",
        DashboardEnrollmentPlan,
        BASE_COLUMNS
        + [
            DashboardColumn("contract_count", "合同例数", "int"),
            DashboardColumn("screening_count", "已筛选病例数", "int"),
            DashboardColumn("current_enrolled_count", "当前入组数", "int"),
            DashboardColumn("positive_enrolled_count", "阳性入组", "int"),
            DashboardColumn("identified_polyp_count", "识别到息肉数量", "int"),
            DashboardColumn("unidentified_polyp_count", "未识别到息肉数量", "int"),
            DashboardColumn("whole_colon_completed_count", "全结肠完成数量", "int"),
            DashboardColumn("whole_colon_incomplete_count", "未全结肠完成数量", "int"),
            DashboardColumn("sigmoid_unidentified_count", "未识别出乙状结肠数量", "int"),
            DashboardColumn("next_week_plan_count", "下周计划入组数", "int"),
            DashboardColumn("eligible_count", "符合入组病例", "int"),
            DashboardColumn("enrollment_arrangement", "入组安排"),
            DashboardColumn("notes", "备注"),
        ],
        ["center_id"],
    ),
    "subject-overviews": DashboardKindConfig(
        "subject-overviews",
        "subject-overviews",
        "整体情况表",
        DashboardSubjectOverview,
        BASE_COLUMNS
        + [
            DashboardColumn("screening_no", "筛选号", "text", True),
            DashboardColumn("informed_at", "知情时间", "datetime"),
            DashboardColumn("swallow_time", "吞服时间", "datetime"),
            DashboardColumn("swallow_time_2", "吞服时间2", "datetime"),
            DashboardColumn("gastric_transit_time", "胃转运时间"),
            DashboardColumn("colon_entry_duration", "进入结肠时长"),
            DashboardColumn("capsule_batch_no", "胶囊批次号"),
            DashboardColumn("capsule_serial_no", "胶囊序列号"),
            DashboardColumn("recorder_batch_no", "记录仪批次号"),
            DashboardColumn("recorder_serial_no", "记录仪序列号"),
            DashboardColumn("image_count", "图像数量", "int"),
            DashboardColumn("video_duration", "视频时长"),
            DashboardColumn("colon_work_duration", "结肠工作时间"),
            DashboardColumn("condition_description", "情况描述"),
            DashboardColumn("capsule_excreted_at", "胶囊排出时间", "datetime"),
        ],
        ["center_id", "screening_no"],
    ),
    "device-handovers": DashboardKindConfig(
        "device-handovers",
        "device-handovers",
        "器械交接记录表",
        DashboardDeviceHandover,
        BASE_COLUMNS
        + [
            DashboardColumn("device_name", "器械名称", "text", True),
            DashboardColumn("batch_no", "批次号"),
            DashboardColumn("device_serial_no", "序列号", "text", True),
            DashboardColumn("handed_over_at", "交接日期", "date"),
            DashboardColumn("returned_at", "归还日期", "date"),
            DashboardColumn("handover_status", "交接状态"),
            DashboardColumn("handover_person", "交接人"),
            DashboardColumn("receiver", "接收人"),
            DashboardColumn("notes", "备注"),
        ],
        ["center_id", "device_name", "device_serial_no"],
    ),
    "subject-results": DashboardKindConfig(
        "subject-results",
        "subject-results",
        "受试者结果统计表",
        DashboardSubjectResult,
        BASE_COLUMNS
        + [
            DashboardColumn("reading_no", "阅片号"),
            DashboardColumn("screening_no", "筛选号", "text", True),
            DashboardColumn("enrollment_no", "入组号"),
            DashboardColumn("whole_colon_completed", "全结肠完成判断"),
            DashboardColumn("is_positive", "是否阳性"),
            DashboardColumn("max_polyp_size", "最大息肉大小"),
            DashboardColumn("capsule_polyp_count", "胶囊息肉数", "int"),
            DashboardColumn("colonoscopy_polyp_count", "电子肠镜息肉数", "int"),
            DashboardColumn("matched_polyp_count", "匹配息肉数", "int"),
            DashboardColumn("is_fully_matched", "是否完全匹配"),
            DashboardColumn("max_polyp_matched", "是否匹配最大息肉"),
            DashboardColumn("other_diagnosis", "其他疾病诊断"),
            DashboardColumn("result_notes", "结果备注"),
        ],
        ["center_id", "screening_no"],
    ),
    "clinical-events": DashboardKindConfig(
        "clinical-events",
        "clinical-events",
        "临床事件记录",
        DashboardClinicalEvent,
        BASE_COLUMNS
        + [
            DashboardColumn("event_name", "事件", "text", True),
            DashboardColumn("occurred_at", "发生时间", "datetime"),
            DashboardColumn("event_type", "事件类型"),
            DashboardColumn("severity", "严重程度"),
            DashboardColumn("status", "状态"),
            DashboardColumn("notes", "备注"),
        ],
        ["event_name", "occurred_at"],
    ),
    "device-issues": DashboardKindConfig(
        "device-issues",
        "device-issues",
        "器械问题记录表",
        DashboardDeviceIssue,
        BASE_COLUMNS
        + [
            DashboardColumn("problem_time", "问题时间", "datetime"),
            DashboardColumn("problem_description", "问题描述", "text", True),
            DashboardColumn("is_resolved", "是否解决"),
            DashboardColumn("problem_type", "问题类型"),
            DashboardColumn("center_institution", "中心机构"),
            DashboardColumn("notes", "备注"),
        ],
        ["center_id", "problem_time", "problem_description"],
    ),
    "important-tasks": DashboardKindConfig(
        "important-tasks",
        "important-tasks",
        "重要紧急事项完成",
        DashboardImportantTask,
        BASE_COLUMNS
        + [
            DashboardColumn("title", "事项", "text", True),
            DashboardColumn("owner", "负责人"),
            DashboardColumn("planned_due_date", "计划完成日期", "date"),
            DashboardColumn("actual_completed_date", "实际完成日期", "date"),
            DashboardColumn("status", "状态"),
            DashboardColumn("importance", "重要程度"),
            DashboardColumn("urgency", "紧急程度"),
            DashboardColumn("notes", "备注"),
        ],
        ["title", "planned_due_date"],
    ),
}


def get_config(kind: str) -> DashboardKindConfig:
    try:
        return DASHBOARD_V31_CONFIGS[kind]
    except KeyError as exc:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="dashboard kind not found"
        ) from exc


def get_project_or_404(db: Session, project_id: int) -> Project:
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="project not found")
    return project


def ensure_project_read(access: AccessContext, project_id: int) -> None:
    if not access.can_access_project(project_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Project scope denied")


def ensure_center_scope(
    db: Session, access: AccessContext, project_id: int, center_id: int | None, write: bool
) -> None:
    if center_id is None:
        if write and not (access.is_admin or project_id in access.project_ids):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, detail="Project write scope denied"
            )
        return
    center = db.get(Center, center_id)
    if center is None or center.project_id != project_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="center does not belong to project"
        )
    if not access.can_access_center(center_id, project_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Center scope denied")


def scoped_statement(
    db: Session,
    access: AccessContext,
    config: DashboardKindConfig,
    project_id: int,
    center_id: int | None,
) -> Select:
    get_project_or_404(db, project_id)
    ensure_project_read(access, project_id)
    statement = select(config.model).where(config.model.project_id == project_id)
    if center_id is not None:
        ensure_center_scope(db, access, project_id, center_id, write=False)
        statement = statement.where(config.model.center_id == center_id)
    elif not access.is_admin and project_id not in access.project_ids:
        statement = statement.where(config.model.center_id.in_(access.center_ids))
    return statement.order_by(config.model.id)


def list_records(
    db: Session,
    access: AccessContext,
    kind: str,
    project_id: int,
    center_id: int | None = None,
) -> list[Any]:
    config = get_config(kind)
    return list(db.scalars(scoped_statement(db, access, config, project_id, center_id)))


def get_record(
    db: Session, access: AccessContext, kind: str, record_id: int, write: bool = False
) -> Any:
    config = get_config(kind)
    record = db.get(config.model, record_id)
    if record is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="dashboard record not found"
        )
    ensure_project_read(access, record.project_id)
    ensure_center_scope(db, access, record.project_id, record.center_id, write=write)
    return record


def create_record(db: Session, access: AccessContext, kind: str, values: dict[str, Any]) -> Any:
    config = get_config(kind)
    project_id = values["project_id"]
    get_project_or_404(db, project_id)
    ensure_project_read(access, project_id)
    ensure_center_scope(db, access, project_id, values.get("center_id"), write=True)
    record = config.model(**values)
    db.add(record)
    commit_or_conflict(db, f"{config.title}记录已存在")
    db.refresh(record)
    return record


def update_record(
    db: Session, access: AccessContext, kind: str, record_id: int, values: dict[str, Any]
) -> Any:
    record = get_record(db, access, kind, record_id, write=True)
    project_id = values.get("project_id", record.project_id)
    center_id = values.get("center_id", record.center_id)
    if project_id != record.project_id:
        get_project_or_404(db, project_id)
        ensure_project_read(access, project_id)
    ensure_center_scope(db, access, project_id, center_id, write=True)
    for key, value in values.items():
        setattr(record, key, value)
    commit_or_conflict(db, "dashboard record already exists")
    db.refresh(record)
    return record


def delete_record(db: Session, access: AccessContext, kind: str, record_id: int) -> None:
    record = get_record(db, access, kind, record_id, write=True)
    db.delete(record)
    db.commit()


def commit_or_conflict(db: Session, message: str) -> None:
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=message) from exc


def build_template_workbook(kind: str) -> bytes:
    config = get_config(kind)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = config.title[:31]
    worksheet.append([column.label for column in config.columns])
    worksheet.append([column.key for column in config.columns])
    worksheet.append([column.note for column in config.columns])
    worksheet.freeze_panes = "A4"
    for index, column in enumerate(config.columns, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = max(
            len(column.label) + 6,
            len(column.key) + 4,
            14,
        )
    return workbook_bytes(workbook)


def export_records_workbook(
    db: Session,
    access: AccessContext,
    kind: str,
    project_id: int,
    center_id: int | None = None,
) -> bytes:
    config = get_config(kind)
    records = list_records(db, access, kind, project_id, center_id)
    centers = {
        center.id: center.code
        for center in db.scalars(select(Center).where(Center.project_id == project_id))
    }
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = config.title[:31]
    worksheet.append([column.label for column in config.columns])
    worksheet.append([column.key for column in config.columns])
    for record in records:
        row = []
        for column in config.columns:
            if column.key == "center_code":
                value = centers.get(record.center_id, "")
            else:
                value = getattr(record, column.key, None)
            row.append(value)
        worksheet.append(row)
    worksheet.freeze_panes = "A3"
    for index, column in enumerate(config.columns, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = max(
            len(column.label) + 6,
            14,
        )
    return workbook_bytes(workbook)


def import_records_workbook(
    db: Session,
    access: AccessContext,
    kind: str,
    project_id: int,
    content: bytes,
) -> DashboardV31ImportResultRead:
    config = get_config(kind)
    get_project_or_404(db, project_id)
    ensure_project_read(access, project_id)
    rows, errors = parse_workbook(content, config)
    created_count = 0
    updated_count = 0
    saved_rows: list[dict[str, Any]] = []
    centers_by_code = {
        center.code: center.id
        for center in db.scalars(select(Center).where(Center.project_id == project_id))
    }

    for row_number, values in rows:
        values["project_id"] = project_id
        center_code = clean_text(values.pop("center_code", None))
        if center_code:
            if center_code not in centers_by_code:
                errors.append(
                    DashboardV31ImportErrorRead(
                        row=row_number, field="center_code", message="中心编码不存在"
                    )
                )
                continue
            values["center_id"] = centers_by_code[center_code]
        else:
            values["center_id"] = None
        missing = [
            column.key
            for column in config.columns
            if column.required and not clean_text(values.get(column.key))
        ]
        if missing:
            errors.extend(
                DashboardV31ImportErrorRead(row=row_number, field=field, message="必填字段为空")
                for field in missing
            )
            continue
        try:
            ensure_center_scope(db, access, project_id, values.get("center_id"), write=True)
            record = find_existing_record(db, config, values)
            if record is None:
                record = config.model(**model_values(config, values))
                db.add(record)
                created_count += 1
            else:
                for key, value in model_values(config, values).items():
                    setattr(record, key, value)
                updated_count += 1
            db.flush()
            saved_rows.append({"row": row_number, "id": record.id})
        except (ValueError, IntegrityError) as exc:
            db.rollback()
            errors.append(
                DashboardV31ImportErrorRead(row=row_number, field="row", message=str(exc))
            )
        except HTTPException:
            db.rollback()
            raise
    if errors:
        db.rollback()
    else:
        db.commit()
    return DashboardV31ImportResultRead(
        total_rows=len(rows),
        created_count=created_count if not errors else 0,
        updated_count=updated_count if not errors else 0,
        errors=errors,
        rows=saved_rows if not errors else [],
    )


def parse_workbook(
    content: bytes, config: DashboardKindConfig
) -> tuple[list[tuple[int, dict[str, Any]]], list[DashboardV31ImportErrorRead]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="invalid workbook"
        ) from exc
    worksheet = workbook.active
    key_row = [clean_text(cell.value) for cell in worksheet[2]]
    expected = [column.key for column in config.columns]
    if key_row[: len(expected)] != expected:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="template columns mismatch"
        )
    rows: list[tuple[int, dict[str, Any]]] = []
    errors: list[DashboardV31ImportErrorRead] = []
    column_by_key = {column.key: column for column in config.columns}
    for row_index in range(4, worksheet.max_row + 1):
        raw = {
            key: worksheet.cell(row=row_index, column=index + 1).value
            for index, key in enumerate(expected)
        }
        if all(value in (None, "") for value in raw.values()):
            continue
        values: dict[str, Any] = {}
        for key, value in raw.items():
            try:
                values[key] = coerce_value(value, column_by_key[key].kind)
            except ValueError as exc:
                errors.append(
                    DashboardV31ImportErrorRead(row=row_index, field=key, message=str(exc))
                )
        rows.append((row_index, values))
    return rows, errors


def coerce_value(value: Any, kind: str) -> Any:
    if value in (None, ""):
        return None
    if kind == "int":
        try:
            return int(value)
        except (TypeError, ValueError) as exc:
            raise ValueError("需要整数") from exc
    if kind == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return date.fromisoformat(str(value).strip())
        except ValueError as exc:
            raise ValueError("日期格式应为 YYYY-MM-DD") from exc
    if kind == "datetime":
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        text = str(value).strip()
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError as exc:
            raise ValueError("时间格式应为 YYYY-MM-DD HH:MM") from exc
    return clean_text(value)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def find_existing_record(
    db: Session, config: DashboardKindConfig, values: dict[str, Any]
) -> Any | None:
    record_id = values.get("id")
    if record_id:
        return db.get(config.model, record_id)
    filters = [config.model.project_id == values["project_id"]]
    for key in config.unique_keys:
        filters.append(
            getattr(config.model, key).is_(None)
            if values.get(key) is None
            else getattr(config.model, key) == values.get(key)
        )
    return db.scalar(select(config.model).where(*filters))


def model_values(config: DashboardKindConfig, values: dict[str, Any]) -> dict[str, Any]:
    allowed = {column.key for column in config.columns if column.key not in {"id", "center_code"}}
    return {
        key: value
        for key, value in values.items()
        if key in allowed or key in {"project_id", "center_id"}
    }


def workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_overview(db: Session, access: AccessContext, project_id: int) -> dict[str, Any]:
    get_project_or_404(db, project_id)
    ensure_project_read(access, project_id)
    counts = {
        kind: len(list_records(db, access, kind, project_id)) for kind in DASHBOARD_V31_CONFIGS
    }
    enrollment_rows = list_records(db, access, "enrollment-plans", project_id)
    subject_count = (
        db.scalar(select(func.count(Subject.id)).where(Subject.project_id == project_id)) or 0
    )
    enrollment = {
        "contract_count": sum_int(row.contract_count for row in enrollment_rows),
        "planned_next_week": sum_int(row.next_week_plan_count for row in enrollment_rows),
        "maintained_current_enrolled": sum_int(
            row.current_enrolled_count for row in enrollment_rows
        ),
        "subject_count": subject_count,
    }
    task_status = Counter(
        row.status for row in list_records(db, access, "important-tasks", project_id)
    )
    return {
        "project_id": project_id,
        "counts": counts,
        "enrollment": enrollment,
        "important_task_status": dict(task_status),
        "deviation_warnings": build_warnings(db, access, project_id),
    }


def build_warnings(db: Session, access: AccessContext, project_id: int) -> list[dict[str, Any]]:
    today = date.today()
    soon = today + timedelta(days=7)
    warnings: list[dict[str, Any]] = []
    for milestone in list_records(db, access, "milestones", project_id):
        if milestone.planned_date is None or is_done(milestone.status, milestone.actual_date):
            continue
        level = warning_level(milestone.planned_date, today, soon)
        if level:
            warnings.append(
                {
                    "source": "milestone",
                    "id": milestone.id,
                    "title": milestone.milestone_name,
                    "center_id": milestone.center_id,
                    "planned_date": milestone.planned_date,
                    "status": milestone.status,
                    "warning_level": level,
                }
            )
    for task in list_records(db, access, "important-tasks", project_id):
        if task.planned_due_date is None or is_done(task.status, task.actual_completed_date):
            continue
        level = warning_level(task.planned_due_date, today, soon)
        if level:
            warnings.append(
                {
                    "source": "important_task",
                    "id": task.id,
                    "title": task.title,
                    "center_id": task.center_id,
                    "planned_date": task.planned_due_date,
                    "status": task.status,
                    "warning_level": level,
                }
            )
    return sorted(warnings, key=lambda item: (item["planned_date"], item["source"], item["id"]))


def warning_level(planned_date: date, today: date, soon: date) -> str | None:
    if planned_date < today:
        return "overdue"
    if planned_date <= soon:
        return "due_soon"
    return None


def is_done(status_value: str | None, completed_at: Any | None) -> bool:
    return completed_at is not None or (status_value or "").strip().lower() in DONE_STATUSES


def sum_int(values: Any) -> int:
    return sum(value or 0 for value in values)
