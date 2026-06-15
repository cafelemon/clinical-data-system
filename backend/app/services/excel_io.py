from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from statistics import median
from typing import Any

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.deps import AccessContext
from app.core.clinical_data import (
    DATA_CHECKING,
    DATA_COMPLETE,
    DATA_INCOMPLETE,
    REVIEW_PENDING,
    REVIEW_REJECTED,
)
from app.models import Center, Project, Stage, StageFile, StageTemplate, Subject, SubjectItem
from app.schemas import ExcelImportErrorRead, ExcelImportResultRead
from app.services.clinical_status import (
    build_completeness_summary,
    build_stage_file_statuses,
    required_item_status,
)
from app.services.stage_config import (
    CENTER_FILE_SCOPE,
    ensure_project_stage_config,
    ensure_template_scope,
    phase_template_scope,
    template_scope_for_option,
    validate_template_stage,
)
from app.services.subject_setup import create_default_subject_sections

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class ExcelColumn:
    key: str
    label: str
    required: bool = False
    note: str = ""


@dataclass(frozen=True)
class RowError:
    row: int
    field: str
    message: str


@dataclass(frozen=True)
class ParsedRow:
    row: int
    values: dict[str, Any]


TEMPLATE_COLUMNS: dict[str, list[ExcelColumn]] = {
    "projects": [
        ExcelColumn("code", "项目编码", True, "唯一编码"),
        ExcelColumn("name", "项目名称", True, "必填"),
        ExcelColumn("status", "状态", False, "默认 active"),
        ExcelColumn("description", "备注", False),
    ],
    "centers": [
        ExcelColumn("project_code", "项目编码", True),
        ExcelColumn("code", "中心编码", True),
        ExcelColumn("name", "中心名称", True),
        ExcelColumn("contact_person", "联系人", False),
        ExcelColumn("status", "状态", False, "默认 active"),
        ExcelColumn("description", "备注", False),
    ],
    "subjects": [
        ExcelColumn("project_code", "项目编码", True),
        ExcelColumn("center_code", "中心编码", True),
        ExcelColumn("screening_no", "筛选号", True),
        ExcelColumn("subject_arm", "分组", True, "experimental/control 或 实验组/对照组"),
        ExcelColumn("gender", "性别", False),
        ExcelColumn("age", "年龄", False),
        ExcelColumn("enrolled_at", "入组日期", False, "YYYY-MM-DD"),
        ExcelColumn("informed_at", "知情时间", False, "YYYY-MM-DD HH:MM"),
        ExcelColumn("visit1_date", "访视1日期", False, "YYYY-MM-DD"),
        ExcelColumn("visit2_date", "访视2日期", False, "YYYY-MM-DD"),
        ExcelColumn("visit3_date", "访视3日期", False, "YYYY-MM-DD"),
        ExcelColumn("visit4_date", "访视4日期", False, "YYYY-MM-DD"),
    ],
    "stage-templates": [
        ExcelColumn("project_code", "项目编码", True),
        ExcelColumn("stage_code", "阶段编码", True),
        ExcelColumn("template_scope", "模板用途", False, "center_file 或 subject_item"),
        ExcelColumn("item_code", "资料项编码", True),
        ExcelColumn("item_name", "资料项名称", True),
        ExcelColumn("required", "是否必填", False, "默认是，可填 true/false 或 是/否"),
        ExcelColumn("sort_order", "排序", False, "默认 0"),
        ExcelColumn("description", "备注", False),
    ],
}

IMPORT_KINDS = frozenset(TEMPLATE_COLUMNS)


def build_template_workbook(kind: str) -> bytes:
    columns = get_template_columns(kind)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = template_sheet_name(kind)
    worksheet.append([column.label for column in columns])
    worksheet.append([column.key for column in columns])
    worksheet.append([column.note for column in columns])
    worksheet.freeze_panes = "A4"
    for index, column in enumerate(columns, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = max(
            len(column.label) + 6,
            len(column.key) + 4,
            14,
        )
    return workbook_bytes(workbook)


def get_template_columns(kind: str) -> list[ExcelColumn]:
    if kind not in TEMPLATE_COLUMNS:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="import template not found",
        )
    return TEMPLATE_COLUMNS[kind]


def import_excel(
    db: Session,
    access: AccessContext,
    kind: str,
    content: bytes,
) -> ExcelImportResultRead:
    if kind == "projects":
        return import_projects(db, access, content)
    if kind == "centers":
        return import_centers(db, access, content)
    if kind == "subjects":
        return import_subjects(db, access, content)
    if kind == "stage-templates":
        return import_stage_templates(db, access, content)
    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import target not found")


def import_projects(
    db: Session,
    access: AccessContext,
    content: bytes,
) -> ExcelImportResultRead:
    rows, errors = parse_workbook(content, "projects")
    seen = ensure_no_duplicate_keys(rows, ["code"], errors)
    planned: list[tuple[ParsedRow, Project | None, dict[str, Any]]] = []
    for row in rows:
        if row.row in seen:
            continue
        code = require_text(row, "code", errors)
        name = require_text(row, "name", errors)
        if not code or not name:
            continue
        project = db.scalar(select(Project).where(Project.code == code))
        if project is None and not access.is_admin:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only admin can create projects",
            )
        if project is not None and not can_write_project(access, project.id):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Project scope denied",
            )
        planned.append(
            (
                row,
                project,
                {
                    "code": code,
                    "name": name,
                    "status": optional_text(row, "status") or "active",
                    "description": optional_text(row, "description"),
                },
            )
        )
    result = persist_import(db, planned, Project, errors, total_rows=len(rows))
    if not result.errors:
        for _, project, values in planned:
            target = project or db.scalar(select(Project).where(Project.code == values["code"]))
            if target is not None:
                ensure_project_stage_config(db, target)
    return result


def import_centers(
    db: Session,
    access: AccessContext,
    content: bytes,
) -> ExcelImportResultRead:
    rows, errors = parse_workbook(content, "centers")
    seen = ensure_no_duplicate_keys(rows, ["project_code", "code"], errors)
    planned: list[tuple[ParsedRow, Center | None, dict[str, Any]]] = []
    for row in rows:
        if row.row in seen:
            continue
        project_code = require_text(row, "project_code", errors)
        code = require_text(row, "code", errors)
        name = require_text(row, "name", errors)
        if not project_code or not code or not name:
            continue
        project = project_by_code(db, row, project_code, errors)
        if project is None:
            continue
        center = db.scalar(
            select(Center).where(Center.project_id == project.id, Center.code == code)
        )
        if center is None:
            if not can_write_project(access, project.id):
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Project write scope denied",
                )
        elif not access.can_access_center(center.id, center.project_id):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Center scope denied")
        planned.append(
            (
                row,
                center,
                {
                    "project_id": project.id,
                    "code": code,
                    "name": name,
                    "contact_person": optional_text(row, "contact_person"),
                    "status": optional_text(row, "status") or "active",
                    "description": optional_text(row, "description"),
                },
            )
        )
    return persist_import(db, planned, Center, errors, total_rows=len(rows))


def import_subjects(
    db: Session,
    access: AccessContext,
    content: bytes,
) -> ExcelImportResultRead:
    rows, errors = parse_workbook(content, "subjects")
    seen = ensure_no_duplicate_keys(rows, ["project_code", "center_code", "screening_no"], errors)
    planned: list[tuple[ParsedRow, Subject | None, dict[str, Any]]] = []
    for row in rows:
        if row.row in seen:
            continue
        project_code = require_text(row, "project_code", errors)
        center_code = require_text(row, "center_code", errors)
        screening_no = require_text(row, "screening_no", errors)
        if not project_code or not center_code or not screening_no:
            continue
        project = project_by_code(db, row, project_code, errors)
        if project is None:
            continue
        subject_arm = subject_arm_value(row, errors)
        if subject_arm is None:
            continue
        age = optional_int(row, "age", errors)
        enrolled_at = optional_date(row, "enrolled_at", errors)
        informed_at = optional_datetime(row, "informed_at", errors)
        center = center_by_code(db, row, project, center_code, errors)
        if center is None:
            continue
        if not access.can_access_center(center.id, project.id):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Center scope denied")
        subject = db.scalar(
            select(Subject).where(
                Subject.project_id == project.id,
                Subject.center_id == center.id,
                Subject.screening_no == screening_no,
            )
        )
        planned.append(
            (
                row,
                subject,
                {
                    "project_id": project.id,
                    "center_id": center.id,
                    "screening_no": screening_no,
                    "subject_arm": subject_arm,
                    "gender": optional_text(row, "gender"),
                    "age": age,
                    "enrolled_at": enrolled_at,
                    "informed_at": informed_at,
                    "visit1_date": optional_date(row, "visit1_date", errors),
                    "visit2_date": optional_date(row, "visit2_date", errors),
                    "visit3_date": optional_date(row, "visit3_date", errors),
                    "visit4_date": optional_date(row, "visit4_date", errors),
                },
            )
        )
    return persist_subject_import(db, access, planned, errors, total_rows=len(rows))


def import_stage_templates(
    db: Session,
    access: AccessContext,
    content: bytes,
) -> ExcelImportResultRead:
    rows, errors = parse_workbook(content, "stage-templates")
    seen = ensure_no_duplicate_keys(rows, ["project_code", "stage_code", "item_code"], errors)
    planned: list[tuple[ParsedRow, StageTemplate | None, dict[str, Any]]] = []
    for row in rows:
        if row.row in seen:
            continue
        project_code = require_text(row, "project_code", errors)
        stage_code = require_text(row, "stage_code", errors)
        item_code = require_text(row, "item_code", errors)
        item_name = require_text(row, "item_name", errors)
        if not project_code or not stage_code or not item_code or not item_name:
            continue
        project = project_by_code(db, row, project_code, errors)
        if project is None:
            continue
        if not can_write_project(access, project.id):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Project write scope denied",
            )
        ensure_project_stage_config(db, project.id)
        stage = db.scalar(
            select(Stage).where(Stage.project_id == project.id, Stage.code == stage_code)
        )
        if stage is None:
            errors.append(RowError(row.row, "stage_code", "阶段不存在"))
            continue
        template_scope = optional_text(row, "template_scope")
        if template_scope is None:
            template_scope = (
                phase_template_scope(stage.code)
                if stage.parent_id is None
                else template_scope_for_option(stage.option_code or stage.code)
            )
        try:
            template_scope = ensure_template_scope(template_scope)
            stage = validate_template_stage(stage, template_scope)
        except HTTPException as exc:
            errors.append(RowError(row.row, "template_scope", str(exc.detail)))
            continue
        template = db.scalar(
            select(StageTemplate).where(
                StageTemplate.project_id == project.id,
                StageTemplate.stage_id == stage.id,
                StageTemplate.template_scope == template_scope,
                StageTemplate.item_code == item_code,
            )
        )
        planned.append(
            (
                row,
                template,
                {
                    "project_id": project.id,
                    "stage_id": stage.id,
                    "template_scope": template_scope,
                    "item_code": item_code,
                    "item_name": item_name,
                    "required": optional_bool(row, "required", errors, default=True),
                    "sort_order": optional_int(row, "sort_order", errors, default=0),
                    "description": optional_text(row, "description"),
                },
            )
        )
    return persist_import(db, planned, StageTemplate, errors, total_rows=len(rows))


def build_project_progress_export(
    db: Session,
    access: AccessContext,
    project_id: int | None = None,
) -> bytes:
    projects = visible_projects(db, access, project_id)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "项目进度"
    append_header(
        worksheet,
        [
            "项目编码",
            "项目名称",
            "可见中心数",
            "受试者数",
            "完成案例数",
            "项目累计天数",
            "平均用时/案例",
            "中位数用时",
            "完整性状态",
            "阶段齐全",
            "阶段核查中",
            "阶段不全",
            "受试者齐全",
            "受试者核查中",
            "受试者不全",
        ],
    )
    for project in projects:
        centers = visible_centers(db, access, project.id)
        subjects = subjects_for_centers(db, project.id, [center.id for center in centers])
        completed_subjects = [
            subject for subject in subjects if subject.data_status == DATA_COMPLETE
        ]
        durations = [
            duration
            for subject in completed_subjects
            if (duration := subject_duration_days(subject)) is not None
        ]
        summary = build_completeness_summary(
            db,
            project_ids=scoped_project_ids(access),
            center_ids=scoped_center_ids(access),
            project_id=project.id,
        )
        worksheet.append(
            [
                project.code,
                project.name,
                len(centers),
                len(subjects),
                len(completed_subjects),
                project_days(project),
                round(sum(durations) / len(durations), 1) if durations else 0.0,
                round(float(median(durations)), 1) if durations else 0.0,
                summary.status,
                summary.stage_files[DATA_COMPLETE],
                summary.stage_files[DATA_CHECKING],
                summary.stage_files[DATA_INCOMPLETE],
                summary.subjects[DATA_COMPLETE],
                summary.subjects[DATA_CHECKING],
                summary.subjects[DATA_INCOMPLETE],
            ]
        )
    autosize(worksheet)
    return workbook_bytes(workbook)


def build_center_status_export(
    db: Session,
    access: AccessContext,
    project_id: int | None = None,
    center_id: int | None = None,
) -> bytes:
    centers = visible_centers(db, access, project_id, center_id)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "中心资料状态"
    append_header(
        worksheet,
        [
            "项目编码",
            "项目名称",
            "中心编码",
            "中心名称",
            "受试者数",
            "完成案例数",
            "完成率",
            "完整性状态",
            "阶段齐全",
            "阶段核查中",
            "阶段不全",
            "受试者齐全",
            "受试者核查中",
            "受试者不全",
            "待审核数",
            "驳回数",
        ],
    )
    if center_id is not None and not centers:
        ensure_center_visible(db, access, center_id, project_id)
    for center in centers:
        subjects = subjects_for_centers(db, center.project_id, [center.id])
        completed = sum(1 for subject in subjects if subject.data_status == DATA_COMPLETE)
        summary = build_completeness_summary(
            db,
            project_ids=scoped_project_ids(access),
            center_ids=scoped_center_ids(access),
            project_id=center.project_id,
            center_id=center.id,
        )
        pending_rejected = center_review_counts(db, center.project_id, center.id)
        worksheet.append(
            [
                center.project.code,
                center.project.name,
                center.code,
                center.name,
                len(subjects),
                completed,
                round(completed / len(subjects) * 100, 1) if subjects else 0.0,
                summary.status,
                summary.stage_files[DATA_COMPLETE],
                summary.stage_files[DATA_CHECKING],
                summary.stage_files[DATA_INCOMPLETE],
                summary.subjects[DATA_COMPLETE],
                summary.subjects[DATA_CHECKING],
                summary.subjects[DATA_INCOMPLETE],
                pending_rejected[REVIEW_PENDING],
                pending_rejected[REVIEW_REJECTED],
            ]
        )
    autosize(worksheet)
    return workbook_bytes(workbook)


def build_subject_completeness_export(
    db: Session,
    access: AccessContext,
    project_id: int | None = None,
    center_id: int | None = None,
) -> bytes:
    subjects = visible_subjects(db, access, project_id, center_id)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "受试者完整性"
    append_header(
        worksheet,
        [
            "项目编码",
            "项目名称",
            "中心编码",
            "中心名称",
            "筛选号",
            "性别",
            "年龄",
            "入组日期",
            "知情时间",
            "访视1日期",
            "访视2日期",
            "访视3日期",
            "访视4日期",
            "资料状态",
            "审核状态",
            "首次完成时间",
            "必填项数",
            "齐全项",
            "核查中项",
            "不全项",
            "待审核项",
            "驳回项",
        ],
    )
    for subject in subjects:
        item_statuses = subject_item_counters(db, subject.id)
        worksheet.append(
            [
                subject.project.code,
                subject.project.name,
                subject.center.code,
                subject.center.name,
                subject.screening_no,
                subject.gender,
                subject.age,
                subject.enrolled_at.isoformat() if subject.enrolled_at else "",
                subject.informed_at.isoformat(timespec="minutes") if subject.informed_at else "",
                subject.visit1_date.isoformat() if subject.visit1_date else "",
                subject.visit2_date.isoformat() if subject.visit2_date else "",
                subject.visit3_date.isoformat() if subject.visit3_date else "",
                subject.visit4_date.isoformat() if subject.visit4_date else "",
                subject.data_status,
                subject.review_status,
                subject.completed_at.isoformat() if subject.completed_at else "",
                item_statuses["required"],
                item_statuses[DATA_COMPLETE],
                item_statuses[DATA_CHECKING],
                item_statuses[DATA_INCOMPLETE],
                item_statuses[REVIEW_PENDING],
                item_statuses[REVIEW_REJECTED],
            ]
        )
    autosize(worksheet)
    return workbook_bytes(workbook)


def build_missing_items_export(
    db: Session,
    access: AccessContext,
    project_id: int | None = None,
    center_id: int | None = None,
) -> bytes:
    centers = visible_centers(db, access, project_id, center_id)
    subjects = visible_subjects(db, access, project_id, center_id)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "缺失项清单"
    append_header(
        worksheet,
        [
            "对象类型",
            "项目编码",
            "项目名称",
            "中心编码",
            "中心名称",
            "阶段/受试者",
            "资料项编码",
            "资料项名称",
            "上传状态",
            "审核状态",
            "完整性状态",
        ],
    )
    append_stage_missing_rows(db, worksheet, centers)
    append_subject_missing_rows(db, worksheet, subjects)
    autosize(worksheet)
    return workbook_bytes(workbook)


def parse_workbook(content: bytes, kind: str) -> tuple[list[ParsedRow], list[RowError]]:
    errors: list[RowError] = []
    columns = get_template_columns(kind)
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        return [], [RowError(1, "file", "无法读取 Excel 文件")]
    worksheet = workbook.active
    header_row, mapping = locate_header(worksheet, columns)
    if header_row is None:
        return [], [RowError(1, "header", "未找到模板字段行")]
    for column in columns:
        if column.required and column.key not in mapping:
            errors.append(RowError(header_row, column.key, "缺少必填列"))
    rows: list[ParsedRow] = []
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        values = {
            key: worksheet.cell(row=row_number, column=column_number).value
            for key, column_number in mapping.items()
        }
        if all(is_empty(value) for value in values.values()):
            continue
        rows.append(ParsedRow(row_number, values))
    return rows, errors


def locate_header(
    worksheet: Worksheet,
    columns: list[ExcelColumn],
) -> tuple[int | None, dict[str, int]]:
    aliases = {column.key: column.key for column in columns}
    aliases.update({column.label: column.key for column in columns})
    required_keys = {column.key for column in columns if column.required}
    max_header_row = min(worksheet.max_row, 6)
    for row_number in range(1, max_header_row + 1):
        mapping: dict[str, int] = {}
        for column_number in range(1, worksheet.max_column + 1):
            header = normalize_text(worksheet.cell(row=row_number, column=column_number).value)
            if header in aliases:
                mapping[aliases[header]] = column_number
        if required_keys.issubset(mapping):
            return row_number, mapping
    return None, {}


def ensure_no_duplicate_keys(
    rows: list[ParsedRow],
    keys: list[str],
    errors: list[RowError],
) -> set[int]:
    seen: dict[tuple[str, ...], int] = {}
    duplicate_rows: set[int] = set()
    for row in rows:
        values = tuple(normalize_text(row.values.get(key)) for key in keys)
        if any(value == "" for value in values):
            continue
        if values in seen:
            duplicate_rows.add(row.row)
            errors.append(RowError(row.row, ",".join(keys), f"与第 {seen[values]} 行业务键重复"))
        else:
            seen[values] = row.row
    return duplicate_rows


def persist_import(
    db: Session,
    planned: list[tuple[ParsedRow, Any | None, dict[str, Any]]],
    model: type[Any],
    errors: list[RowError],
    total_rows: int,
) -> ExcelImportResultRead:
    if errors:
        return import_result(total_rows, 0, 0, errors)
    created = 0
    updated = 0
    try:
        for _, instance, values in planned:
            if instance is None:
                db.add(model(**values))
                created += 1
                continue
            for field, value in values.items():
                setattr(instance, field, value)
            updated += 1
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Excel import conflicts with existing data",
        ) from exc
    return import_result(total_rows, created, updated, [])


def persist_subject_import(
    db: Session,
    access: AccessContext,
    planned: list[tuple[ParsedRow, Subject | None, dict[str, Any]]],
    errors: list[RowError],
    total_rows: int,
) -> ExcelImportResultRead:
    if errors:
        return import_result(total_rows, 0, 0, errors)
    created = 0
    updated = 0
    try:
        for _, subject, values in planned:
            if subject is None:
                subject = Subject(**values, added_by=access.user.id)
                db.add(subject)
                db.flush()
                create_default_subject_sections(db, subject)
                created += 1
                continue
            for field, value in values.items():
                setattr(subject, field, value)
            updated += 1
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Excel import conflicts with existing data",
        ) from exc
    return import_result(total_rows, created, updated, [])


def import_result(
    total_rows: int,
    created: int,
    updated: int,
    errors: list[RowError],
) -> ExcelImportResultRead:
    return ExcelImportResultRead(
        total_rows=total_rows,
        created_count=created,
        updated_count=updated,
        skipped_count=len(errors),
        errors=[
            ExcelImportErrorRead(row=error.row, field=error.field, message=error.message)
            for error in errors
        ],
    )


def require_text(row: ParsedRow, key: str, errors: list[RowError]) -> str | None:
    value = normalize_text(row.values.get(key))
    if value == "":
        errors.append(RowError(row.row, key, "必填"))
        return None
    return value


def optional_text(row: ParsedRow, key: str) -> str | None:
    value = normalize_text(row.values.get(key))
    return value or None


def subject_arm_value(row: ParsedRow, errors: list[RowError]) -> str | None:
    value = normalize_text(row.values.get("subject_arm"))
    if value == "":
        errors.append(RowError(row.row, "subject_arm", "必填"))
        return None
    normalized = value.lower()
    if normalized in {"experimental", "实验组"}:
        return "experimental"
    if normalized in {"control", "对照组"}:
        return "control"
    errors.append(RowError(row.row, "subject_arm", "需为 experimental/control 或 实验组/对照组"))
    return None


def optional_int(
    row: ParsedRow,
    key: str,
    errors: list[RowError],
    default: int | None = None,
) -> int | None:
    value = row.values.get(key)
    if is_empty(value):
        return default
    if isinstance(value, bool):
        errors.append(RowError(row.row, key, "需为整数"))
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = normalize_text(value)
    try:
        return int(text)
    except ValueError:
        errors.append(RowError(row.row, key, "需为整数"))
        return default


def optional_bool(
    row: ParsedRow,
    key: str,
    errors: list[RowError],
    default: bool,
) -> bool:
    value = row.values.get(key)
    if is_empty(value):
        return default
    if isinstance(value, bool):
        return value
    text = normalize_text(value).lower()
    if text in {"true", "1", "yes", "y", "是", "必填"}:
        return True
    if text in {"false", "0", "no", "n", "否", "选填"}:
        return False
    errors.append(RowError(row.row, key, "需为布尔值"))
    return default


def optional_date(
    row: ParsedRow,
    key: str,
    errors: list[RowError],
) -> date | None:
    value = row.values.get(key)
    if is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    try:
        return date.fromisoformat(text)
    except ValueError:
        errors.append(RowError(row.row, key, "日期格式应为 YYYY-MM-DD"))
        return None


def optional_datetime(
    row: ParsedRow,
    key: str,
    errors: list[RowError],
) -> datetime | None:
    value = row.values.get(key)
    if is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = normalize_text(value)
    try:
        return datetime.fromisoformat(text).replace(second=0, microsecond=0)
    except ValueError:
        errors.append(RowError(row.row, key, "时间格式应为 YYYY-MM-DD HH:MM"))
        return None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def is_empty(value: Any) -> bool:
    return normalize_text(value) == ""


def project_by_code(
    db: Session,
    row: ParsedRow,
    project_code: str,
    errors: list[RowError],
) -> Project | None:
    project = db.scalar(select(Project).where(Project.code == project_code))
    if project is None:
        errors.append(RowError(row.row, "project_code", "项目不存在"))
    return project


def center_by_code(
    db: Session,
    row: ParsedRow,
    project: Project,
    center_code: str,
    errors: list[RowError],
) -> Center | None:
    center = db.scalar(
        select(Center).where(Center.project_id == project.id, Center.code == center_code)
    )
    if center is None:
        errors.append(RowError(row.row, "center_code", "中心不存在"))
    return center


def can_write_project(access: AccessContext, project_id: int) -> bool:
    return access.is_admin or project_id in access.project_ids


def ensure_project_visible(db: Session, access: AccessContext, project_id: int) -> Project:
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="project not found")
    if not access.can_access_project(project_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Project scope denied")
    return project


def ensure_center_visible(
    db: Session,
    access: AccessContext,
    center_id: int,
    project_id: int | None = None,
) -> Center:
    center = db.get(Center, center_id)
    if center is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="center not found")
    if project_id is not None and center.project_id != project_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="center does not belong to project",
        )
    if not access.can_access_center(center.id, center.project_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Center scope denied")
    return center


def visible_projects(
    db: Session,
    access: AccessContext,
    project_id: int | None = None,
) -> list[Project]:
    statement = select(Project).order_by(Project.id)
    if project_id is not None:
        ensure_project_visible(db, access, project_id)
        statement = statement.where(Project.id == project_id)
    elif not access.is_admin:
        project_ids = access.project_ids | access.center_project_ids
        if not project_ids:
            return []
        statement = statement.where(Project.id.in_(project_ids))
    return list(db.scalars(statement))


def visible_centers(
    db: Session,
    access: AccessContext,
    project_id: int | None = None,
    center_id: int | None = None,
) -> list[Center]:
    if project_id is not None:
        ensure_project_visible(db, access, project_id)
    if center_id is not None:
        ensure_center_visible(db, access, center_id, project_id)
    statement = select(Center).join(Project).order_by(Center.project_id, Center.id)
    if project_id is not None:
        statement = statement.where(Center.project_id == project_id)
    if center_id is not None:
        statement = statement.where(Center.id == center_id)
    if not access.is_admin:
        conditions = []
        if access.project_ids:
            conditions.append(Center.project_id.in_(access.project_ids))
        if access.center_ids:
            conditions.append(Center.id.in_(access.center_ids))
        if not conditions:
            return []
        statement = statement.where(or_(*conditions))
    return list(db.scalars(statement))


def visible_subjects(
    db: Session,
    access: AccessContext,
    project_id: int | None = None,
    center_id: int | None = None,
) -> list[Subject]:
    center_ids = [center.id for center in visible_centers(db, access, project_id, center_id)]
    if not center_ids:
        return []
    statement = (
        select(Subject)
        .where(Subject.center_id.in_(center_ids))
        .order_by(Subject.project_id, Subject.center_id, Subject.id)
    )
    if project_id is not None:
        statement = statement.where(Subject.project_id == project_id)
    return list(db.scalars(statement))


def subjects_for_centers(db: Session, project_id: int, center_ids: list[int]) -> list[Subject]:
    if not center_ids:
        return []
    return list(
        db.scalars(
            select(Subject)
            .where(Subject.project_id == project_id, Subject.center_id.in_(center_ids))
            .order_by(Subject.center_id, Subject.id)
        )
    )


def scoped_project_ids(access: AccessContext) -> set[int] | None:
    if access.is_admin:
        return None
    return access.project_ids


def scoped_center_ids(access: AccessContext) -> set[int] | None:
    if access.is_admin:
        return None
    return access.center_ids


def project_days(project: Project) -> int:
    return max((date.today() - project.created_at.date()).days + 1, 1)


def subject_duration_days(subject: Subject) -> int | None:
    if subject.completed_at is None:
        return None
    start = subject.enrolled_at or subject.created_at.date()
    return max((subject.completed_at.date() - start).days, 0)


def center_review_counts(db: Session, project_id: int, center_id: int) -> Counter[str]:
    counts: Counter[str] = Counter()
    for stage_status in build_stage_file_statuses(
        db,
        project_id=project_id,
        center_id=center_id,
    ):
        if stage_status.review_status in {REVIEW_PENDING, REVIEW_REJECTED}:
            counts[stage_status.review_status] += 1
    subject_statuses = db.execute(
        select(SubjectItem.review_status)
        .join(Subject, SubjectItem.subject_id == Subject.id)
        .where(
            Subject.project_id == project_id,
            Subject.center_id == center_id,
            SubjectItem.required.is_(True),
        )
    )
    for (review_status,) in subject_statuses:
        if review_status in {REVIEW_PENDING, REVIEW_REJECTED}:
            counts[review_status] += 1
    return counts


def subject_item_counters(db: Session, subject_id: int) -> Counter[str]:
    counts: Counter[str] = Counter()
    items = list(
        db.scalars(
            select(SubjectItem)
            .where(SubjectItem.subject_id == subject_id, SubjectItem.required.is_(True))
            .order_by(SubjectItem.id)
        )
    )
    counts["required"] = len(items)
    for item in items:
        status_value = required_item_status(
            item.upload_status,
            item.review_status,
            item.required,
        )
        counts[status_value] += 1
        if item.review_status in {REVIEW_PENDING, REVIEW_REJECTED}:
            counts[item.review_status] += 1
    return counts


def append_stage_missing_rows(db: Session, worksheet: Worksheet, centers: list[Center]) -> None:
    for center in centers:
        templates = list(
            db.scalars(
                select(StageTemplate)
                .where(
                    StageTemplate.project_id == center.project_id,
                    StageTemplate.template_scope == CENTER_FILE_SCOPE,
                    StageTemplate.required.is_(True),
                )
                .order_by(StageTemplate.stage_id, StageTemplate.sort_order, StageTemplate.id)
            )
        )
        stage_by_id = {
            stage.id: stage
            for stage in db.scalars(select(Stage).where(Stage.project_id == center.project_id))
        }
        stage_files = {
            stage_file.stage_template_id: stage_file
            for stage_file in db.scalars(
                select(StageFile).where(
                    StageFile.project_id == center.project_id,
                    StageFile.center_id == center.id,
                    StageFile.stage_template_id.is_not(None),
                )
            )
        }
        for template in templates:
            stage_file = stage_files.get(template.id)
            upload_status = stage_file.upload_status if stage_file else "not_uploaded"
            review_status = stage_file.review_status if stage_file else "unreviewed"
            completeness = required_item_status(upload_status, review_status, True)
            if completeness == DATA_COMPLETE:
                continue
            stage = stage_by_id.get(template.stage_id)
            worksheet.append(
                [
                    "阶段资料",
                    center.project.code,
                    center.project.name,
                    center.code,
                    center.name,
                    stage.name if stage else "",
                    template.item_code,
                    template.item_name,
                    upload_status,
                    review_status,
                    completeness,
                ]
            )


def append_subject_missing_rows(
    db: Session,
    worksheet: Worksheet,
    subjects: list[Subject],
) -> None:
    for subject in subjects:
        items = list(
            db.scalars(
                select(SubjectItem)
                .where(SubjectItem.subject_id == subject.id, SubjectItem.required.is_(True))
                .order_by(SubjectItem.sort_order, SubjectItem.id)
            )
        )
        for item in items:
            completeness = required_item_status(
                item.upload_status,
                item.review_status,
                item.required,
            )
            if completeness == DATA_COMPLETE:
                continue
            worksheet.append(
                [
                    "受试者数据项",
                    subject.project.code,
                    subject.project.name,
                    subject.center.code,
                    subject.center.name,
                    subject.screening_no,
                    item.item_code,
                    item.item_name,
                    item.upload_status,
                    item.review_status,
                    completeness,
                ]
            )


def append_header(worksheet: Worksheet, headers: list[str]) -> None:
    worksheet.append(headers)


def autosize(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        width = max(len(normalize_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 12), 36)


def workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def template_sheet_name(kind: str) -> str:
    return {
        "projects": "项目导入",
        "centers": "中心导入",
        "subjects": "受试者导入",
        "stage-templates": "阶段资料模板导入",
    }[kind]
